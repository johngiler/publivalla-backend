"""
Plantillas Excel para importación masiva de empresas cliente y pedidos históricos.

Los archivos generados se entregan a tenants que migran datos legacy (sin PDFs,
comprobantes ni estados del flujo). El comando de importación se implementará aparte.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill("solid", fgColor="FFF4F4F5")
_REQUIRED_FILL = PatternFill("solid", fgColor="FFFEF3C7")
_NOTE_FONT = Font(italic=True, color="FF52525B")
_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _autosize_columns(ws, *, max_width: int = 52, scan_rows: int = 40) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        best = 12
        for row in range(1, min(ws.max_row, scan_rows) + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            best = min(max(best, len(str(value)) + 2), max_width)
        ws.column_dimensions[letter].width = best


def _write_headers(ws, headers: list[tuple[str, bool]], *, row: int = 1) -> None:
    """headers: (label, required)."""
    for col, (label, required) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _REQUIRED_FILL if required else _HEADER_FILL
        cell.alignment = _WRAP


def _add_instructions_sheet(wb: Workbook, title: str, lines: list[str]) -> None:
    ws = wb.create_sheet(title)
    ws.cell(row=1, column=1, value="Instrucciones").font = Font(bold=True, size=12)
    for i, line in enumerate(lines, start=3):
        cell = ws.cell(row=i, column=1, value=line)
        cell.alignment = _WRAP
        cell.font = _NOTE_FONT if line.startswith("·") or line.startswith("-") else Font()
    ws.column_dimensions["A"].width = 96
    for row in range(3, 3 + len(lines)):
        ws.row_dimensions[row].height = 28


CLIENT_HEADERS: list[tuple[str, bool]] = [
    ("empresa_nombre", True),
    ("email", True),
    ("rif", False),
    ("clave_empresa", False),
    ("contacto", False),
    ("representante_legal", False),
    ("cedula_representante", False),
    ("telefono", False),
    ("direccion", False),
    ("ciudad", False),
    ("notas", False),
    ("estado", False),
]

CLIENT_EXAMPLE_ROW = [
    "Distribuidora Los Andes, C.A.",
    "compras@distlosandes.com.ve",
    "J-30845219-6",
    "LOS-ANDES",
    "Ricardo Salazar",
    "Ricardo Salazar",
    "V-12.345.678",
    "+58 212 482-9130",
    "Av. Francisco de Miranda, Torre Empresarial, piso 4",
    "Caracas",
    "Cliente histórico migrado",
    "active",
]

CLIENT_INSTRUCTIONS = [
    "Plantilla para cargar empresas cliente en Publivalla.",
    "",
    "Columnas obligatorias (fondo amarillo):",
    "· empresa_nombre — Razón social o nombre comercial (máx. 255 caracteres).",
    "· email — Correo de contacto de la empresa (único recomendado por fila).",
    "",
    "Columnas opcionales:",
    "· rif — Identificación fiscal (máx. 32 caracteres). Si se indica, debe ser único dentro del workspace.",
    "· clave_empresa — Código interno suyo para enlazar pedidos (máx. 64 caracteres). Útil si aún no tienen RIF.",
    "· contacto — Persona de contacto operativo.",
    "· representante_legal — Nombre del firmante en documentos.",
    "· cedula_representante — Cédula del representante (ej. V-17.311.805).",
    "· telefono, direccion, ciudad, notas — Datos de contacto y observaciones.",
    "· estado — active (activo) o suspended (suspendido). Si se omite, se asume active.",
    "",
    "No incluir en esta plantilla:",
    "- Imagen de portada, usuarios de acceso ni workspace (se asigna al importar).",
    "",
    "Consejos:",
    "- Una fila = una empresa.",
    "- Complete rif o clave_empresa (o ambos) para poder referenciar la empresa en la plantilla de pedidos.",
    "- La fila 2 es solo un ejemplo; bórrela o sustitúyala antes de importar.",
]

ORDER_HEADERS: list[tuple[str, bool]] = [
    ("referencia_pedido", True),
    ("cliente_rif", False),
    ("cliente_email", False),
    ("cliente_clave", False),
    ("fecha_pedido", True),
    ("marca", False),
    ("campana", False),
    ("descripcion_actividad", False),
    ("notas", False),
    ("instagram", False),
]

ORDER_EXAMPLE_ROW = [
    "PED-2024-001",
    "J-30845219-6",
    "compras@distlosandes.com.ve",
    "LOS-ANDES",
    "2024-06-15",
    "Marca ejemplo",
    "Campaña verano 2024",
    "Activación en pasillo principal",
    "",
    "marca_ejemplo",
]

ORDER_LINE_HEADERS: list[tuple[str, bool]] = [
    ("referencia_pedido", True),
    ("codigo_toma", True),
    ("meses_reservados", True),
    ("subtotal_acordado_usd", False),
]

ORDER_LINE_EXAMPLE_ROW = [
    "PED-2024-001",
    "SPA-T4A",
    "2024-06, 2024-07, 2024-08",
    "900.00",
]

ORDER_INSTRUCTIONS = [
    "Plantilla para cargar pedidos históricos (reservas ya cerradas o en curso).",
    "",
    "Hoja «Pedidos» — una fila por pedido:",
    "· referencia_pedido (obligatorio) — Código suyo para agrupar líneas (ej. PED-2024-001). No es el código automático de Publivalla.",
    "· cliente_rif, cliente_email o cliente_clave (al menos uno) — Debe coincidir con una empresa de la plantilla de clientes.",
    "· fecha_pedido (obligatorio) — Fecha del pedido o de la reserva (AAAA-MM-DD).",
    "· marca, campana, descripcion_actividad, notas, instagram — Información comercial opcional.",
    "",
    "Hoja «Lineas» — una fila por toma reservada:",
    "· referencia_pedido (obligatorio) — Mismo valor que en la hoja Pedidos.",
    "· codigo_toma (obligatorio) — Código de la toma en Publivalla (ej. SPA-T4A). Debe existir en el catálogo del workspace.",
    "· meses_reservados (obligatorio) — Meses de calendario reservados, separados por coma (ej. 2024-06, 2024-07).",
    "· subtotal_acordado_usd (opcional) — Importe acordado sin IVA para esa línea. Si se omite, se calculará según el precio del catálogo.",
    "",
    "No incluir en esta plantilla:",
    "- Estados del flujo, PDFs, hojas firmadas, comprobantes, facturas ni permisos.",
    "",
    "Consejos:",
    "- Un pedido con varias tomas = varias filas en «Lineas» con la misma referencia_pedido.",
    "- Los meses no tienen que ser consecutivos (ej. 2024-06, 2024-09).",
    "- Las filas 2 de cada hoja son solo ejemplos; bórrelas o sustitúyalas antes de importar.",
]


def build_clients_import_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas"
    _write_headers(ws, CLIENT_HEADERS)
    for col, value in enumerate(CLIENT_EXAMPLE_ROW, start=1):
        ws.cell(row=2, column=col, value=value)
    _autosize_columns(ws)
    _add_instructions_sheet(wb, "Instrucciones", CLIENT_INSTRUCTIONS)
    return wb


def build_orders_import_workbook() -> Workbook:
    wb = Workbook()
    ws_o = wb.active
    ws_o.title = "Pedidos"
    _write_headers(ws_o, ORDER_HEADERS)
    for col, value in enumerate(ORDER_EXAMPLE_ROW, start=1):
        ws_o.cell(row=2, column=col, value=value)
    _autosize_columns(ws_o)

    ws_l = wb.create_sheet("Lineas")
    _write_headers(ws_l, ORDER_LINE_HEADERS)
    for col, value in enumerate(ORDER_LINE_EXAMPLE_ROW, start=1):
        ws_l.cell(row=2, column=col, value=value)
    _autosize_columns(ws_l)

    _add_instructions_sheet(wb, "Instrucciones", ORDER_INSTRUCTIONS)
    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def default_import_templates_dir() -> Path:
    return Path(__file__).resolve().parents[3] / "data" / "import_templates"


def write_import_templates(
    output_dir: Path | None = None,
) -> tuple[Path, Path]:
    """Escribe las plantillas .xlsx en disco y devuelve las rutas generadas."""
    out = output_dir or default_import_templates_dir()
    out.mkdir(parents=True, exist_ok=True)

    clients_path = out / "plantilla_empresas_clientes.xlsx"
    orders_path = out / "plantilla_pedidos_historicos.xlsx"

    build_clients_import_workbook().save(clients_path)
    build_orders_import_workbook().save(orders_path)
    return clients_path, orders_path
